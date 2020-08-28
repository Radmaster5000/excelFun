##########################################################################
#                     Eastings and Northings grabber                     #
#      Requires a list of postcodes in the first column starting in      #
#      cell 2. Postcodes need to be formatted properly. Sheet needs      #
#      two further columns; the first titled 'Eastings' and the second   #
#      titled 'Northings'. Program will find the Eastings and Northings  #
#      of the given postcodes and print them in the adjacent empty       #
#                                   cells.                               #
#                                                                        #
##########################################################################

import openpyxl, os, time
from selenium import webdriver # Make sure the 'geckodriver' executable needs to be in PATH message isn't showing up
from introScreen import *
from checkSpeed import *
from openpyxl.utils.exceptions import InvalidFileException
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException

# Add a useage message. e.g. Usage: EastNorth.py 'testBook.xlsx' [sheetname]

#xlFileName = argv1 # get excel file name
#sheetName = argv2 # get sheet name if applicable
postcodes = [] # start with an empty list
eastings = []
northings = []

start = False

# Function for scraping the doogal.co.uk website

def doogalScrape(postcodes, sheet):
	try:
		browser = webdriver.Firefox()
	except WebDriverException:
		print('try typing the following into terminal:')
		print('export PATH=$PATH:/Users/Radmaster5000/Desktop')
		quit()

	browser.get('http://www.doogal.co.uk')
	row = 2
	eastingColumn = 2
	northingColumn = 3

	for postcode in postcodes:
		elem = browser.find_element_by_css_selector('#search')
		elem.send_keys(postcode)
		elem.submit()
		# Page needs to load before it can successfully find the CSS elements for the Easting and Northing
		time.sleep(speed)
		
		try:
			easting = browser.find_element_by_css_selector('div.row:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > table:nth-child(3) > tbody:nth-child(1) > tr:nth-child(3) > td:nth-child(3)').text
		except NoSuchElementException:
			easting = "Error with postcode"

		try:
			northing = browser.find_element_by_css_selector('div.row:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > table:nth-child(3) > tbody:nth-child(1) > tr:nth-child(4) > td:nth-child(3)').text
		except NoSuchElementException:
			northing = "Error with postcode"

		# Print the eastings and northings to the adjacent cells on the spreadsheet	
		sheet.cell(row = row, column = eastingColumn, value = easting)
		sheet.cell(row = row, column = northingColumn, value = northing)	
	
		row += 1

	return 

print(start)

intro()

status = input()

if(status == '\n'):
	start = True

print(start)

xlFileName = input('Enter Excel Workbook name (including the extension):\n') # get excel file name
sheetName = input('Enter Worksheet name:\n') # get sheet name if applicable
speed = int(checkSpeed())

# if no sheetname is inputted, default to standard sheet name.
#if(sheetName == NULL):
#	sheetName = 'Sheet1'

# check if user wants the original spreadsheet overwritten:
overwriteYN = input('Would you like to overwrite the original workbook? Y/N\n')
if (overwriteYN == 'Y' or overwriteYN == 'y'):
	saveFileName = xlFileName
else:
	saveFileName = 'EastNorth.xlsx'


try:
	wb = openpyxl.load_workbook(xlFileName)
except InvalidFileException:
	print("don't forget the extension")
	quit()
except FileNotFoundError:
	print('Please make sure the excel file is in the current directory')
	quit()

sheet = wb[sheetName]


# Load the postcodes into the postcodes list from the workbook
# First get the number of postcodes (assuming the postcodes are in column 'A')
numOfPostcodes = len(sheet['A'])

# Loop through the cells in the first column, starting from cell 2 and add their values to the postcode list
for listItem in range(2, numOfPostcodes+1): 
	postcodes.append(sheet.cell(row=listItem, column=1).value)

print("Running... Please wait...")
# Open the Doogal website then loop through each postcode, scraping the eastings and northings for each one
# *** Possibly just print from the spreadsheet using print(sheet[A2'].value) ***
# Print the eastings and northings from the lists into the correct columns on the spreadsheet
doogalScrape(postcodes, sheet)

wb.save(saveFileName)
wb.close()

print("All done!")
print(":)")
